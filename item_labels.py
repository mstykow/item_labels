#! python3
# Program to create item labels on Avery 5160 from QuickBooks export file.

# Install these modules before first time running script.
import labels, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from reportlab.graphics import shapes
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFont, stringWidth

# Python standard modules.
import sys, os, warnings, re, math
warnings.simplefilter("ignore")

# Letter size sheet (215.9mm x 279.4mm) with Avery 5160 labels.
# 3 columns and 10 rows with labels of size 101.6mm x 50.8mm and a 2mm corner radius.
# The margins are as per specifications from the manufacturer.
specs = labels.Specification(215.9, 279.4, 3, 10, 66.8, 25.4, corner_radius=2,
                             left_padding=1, top_padding=1, bottom_padding=1, right_padding=1, padding_radius=0,
                             left_margin=4.8, column_gap=3, top_margin=12.7, row_gap=0)


# Get the path to the demos directory.
base_path = os.path.dirname(sys.argv[0])
os.chdir(base_path)

# Add some fonts.
registerFont(TTFont('Calibri', os.path.join(base_path, 'calibri.ttf')))
registerFont(TTFont('Calibrib', os.path.join(base_path, 'calibrib.ttf')))
registerFont(TTFont('Calibrii', os.path.join(base_path, 'calibrii.ttf')))

# Opens sheet with the data, determines the last row, and creates save file name.
def export_file_processor(exportFile):
    while os.path.isfile(exportFile) == False:
        print("Filename does not exist. Please try again.")
        exportFile = input()    
    print("Processing...")
    workbook = openpyxl.load_workbook(exportFile)
    worksheet = workbook.get_sheet_by_name('Sheet1')
    maxRow = worksheet.max_row
    filename = os.path.splitext(exportFile)[0] + '.pdf'
    return workbook, worksheet, maxRow, filename

# Creates new sheet with a given list of fields inside a workbook.
def make_import_sheet(workbook, name, fieldList):
    worksheet = workbook.create_sheet(title = name)
    for i in range(len(fieldList)):
        worksheet.cell(row = 1, column = i + 1).value = list(fieldList.keys())[i]
    return worksheet

# Finds the first row containing data in column 'B' of a sheet.
def find_data(sheet):
    startRow = 1
    for cell in list(sheet.columns)[1]:
        if cell.value == None:
            startRow += 1
        else:
            break
    return startRow

# Finds columns corresponding to each cell in a list and returns assignment as a dictionary.
def find_columns(listOfCells, sheet):
    dictionary = {}
    for cell in list(sheet.rows)[startRow - 1]:
        if cell.value in listOfCells:
            dictionary[cell.value] = cell.column
        else:
            continue
    return dictionary

# Writes source cells in sheet1 to target cells in sheet2.
def source_to_target(sheet1, start, end, sheet2, sourceDict, targetDict):
    for i in range(2, end - (start - 1) + 1):
        for key in list(sourceDict.keys()):
            sheet2[targetDict[key] + str(i)].value = sheet1[sourceDict[key] + \
                                                                     str(start + i - 1)].value

# Writes a value into column B depending on a regex found in a cell in column A.
def cert_status(sheet, A, B, dictionary):
    for memo in list(sheet.columns)[column_index_from_string(A) - 1][1:]:
        mo = memoRegex.search(str(memo.value))
        if mo:
            sheet[B + str(memo.row)].value = dictionary.get(mo.group(), None)
        else:
            continue

# Create a function to draw each label. This will be given the ReportLab drawing
# object to draw on, the dimensions (NB. these will be in points, the unit
# ReportLab uses) of the label, and the name to put on the tag.
def write_name(label, width, height, row):
    # Write the item code.
    label.add(shapes.String(width/2.0, height-13, str(row[4].value), fontName='Calibrib', fontSize=11, textAnchor="middle"))
    # Write the invoice number and quantity.
    label.add(shapes.String(5, height/2.0 + 2, 'Invoice #: ' + str(row[0].value), fontName='Calibri', fontSize=11))
    qty_width = stringWidth('Quantity: ' + str(row[3].value), "Calibri", 11) + 5
    label.add(shapes.String(width - qty_width, height/2.0 + 2, 'Quantity: ' + str(row[3].value), fontName='Calibri', fontSize=11))
    # Write certification level.
    label.add(shapes.String(5, height/2.0 - 14, str(row[5].value), fontName='Calibri', fontSize=10))
    # Write website and location.
    label.add(shapes.String(5, 7, 'RasaCreekFarm.com', fontName='Calibri', fontSize=10))
    loc_width = stringWidth('Lumby, BC', "Calibri", 10) + 5
    label.add(shapes.String(width - loc_width, 7, 'Lumby, BC', fontName='Calibri', fontSize=10))

# Given two numbers, this function creates a tuple of tuples.
def numpair_tuples(top, bottom):
    numList = []
    topRows = math.ceil(top/3)
    bottomRows = math.floor(bottom/3)
    if top % 3 == 0:
        for i in range(1, topRows + 1):
            for j in range(1, 4):
               numList.append((i, j))
    else:
        for i in range(1, topRows):
            for j in range(1, 4):
                numList.append((i, j))
        for k in range(1, top % 3 + 1):
            numList.append((topRows, k))
    if bottom % 3 == 0:
        for i in range(10, 10 - bottomRows, -1):
            for j in range (1, 4):
                numList.append((i, j))
    else:
        for i in range(10, 10 - bottomRows, -1):
            for j in range (1, 4):
                numList.append((i, j))
        for k in range(3, 3 - bottom % 3, -1):
            numList.append((10 - bottomRows, k))
    return tuple(numList)

# Creates labels in pdfSheet row by row from xlsSheet if key column is non-empty.
def create_labels(xlsSheet, pdfSheet, column):
    for rowOfCells in list(xlsSheet.rows)[1:]:
        if rowOfCells[column_index_from_string(column)].value:
            pdfSheet.add_label(rowOfCells)
        else:
            continue

# Regex to select content between two enclosing brackets.
memoRegex = re.compile('(?<=\().*?(?=\))')

# Fields required for labels.
reqFields = {
    'Num': 'A',
    'Name': 'B',
    'Memo': 'C',
    'Qty': 'D',
    'Label': 'E',
    'Certification': 'F',
}

# Fields from QuickBooks export.
srcFieldsList = [
    'Num',
    'Name',
    'Memo',
    'Qty',
    'Label',
]

# Certification status options.
certStatus = {
    'cert. organic': 'Certified organic by PACS# 16-608',
    'non-organic': 'Non-certified',
    'cert. organic in BC': 'Certified organic in BC by PACS# 16-608'
}

# Get QuickBooks export file.
print('''Please make sure that the QuickBooks export file is located in the
same directory as this program.''')
print("Please enter the export file's filename (e.g. export.xlsx):")
exportFile = input()
itemsWB, itemsSheet, maxRow, saveFile = export_file_processor(exportFile)
print('How many labels are used at the TOP of the first page?')
topUsed = int(input())
print('How many labels are used at the BOTTOM of the first page?')
bottomUsed = int(input())
print('Processing...')

# Make sheet with required fields.
labelSheet = make_import_sheet(itemsWB, 'Labels', reqFields)

# Find top-left corner of data table.
startRow = find_data(itemsSheet)

# Find columns of source fields.
srcFieldsDict = find_columns(srcFieldsList, itemsSheet)

# Map source fields to required fields and write to labelSheet if 'Label' cell is non-empty.
source_to_target(itemsSheet, startRow, maxRow, labelSheet, srcFieldsDict, reqFields)

# Assign certification status based on memo cell.
reqCol1 = reqFields['Memo']
reqCol2 = reqFields['Certification']
cert_status(labelSheet, reqCol1, reqCol2, certStatus)

# Create the sheet.
itemLabels = labels.Sheet(specs, write_name, border=False)
gridTuple = numpair_tuples(topUsed, bottomUsed)
itemLabels.partial_page(1, gridTuple)

# Create PDF labels row by row from labelSheet.
reqCol3 = reqFields['Label']
create_labels(labelSheet, itemLabels, reqCol3)

# Uncomment the following block to save data to Excel file also.
# This may be useful for debugging the above code.

#try:
#    print('Saving Excel data...')
#    itemsWB.save(exportFile)
#    print('EST import data also saved to sheet "EST" of %s.' % (exportFile))
#except:
#    print('''Could not also save EST import data in %s
#          because it is open in another application.''' % (exportFile))

# Save PDF file and we are done.
try:
    itemLabels.save(saveFile)
    print("{0:d} label(s) output on {1:d} page(s).".format(itemLabels.label_count, itemLabels.page_count))
except:
    print('Could not update PDF file because it is open in another application.')
